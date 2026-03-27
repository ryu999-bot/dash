"""바른손몰 affiliate (제휴업체) analysis endpoints."""
import io
import re
from flask import Blueprint, request, jsonify, send_file
from db.connection import query
from db.queries_affiliate import AFFILIATE_RANKING, AFFILIATE_MONTHLY

bp = Blueprint("affiliate", __name__)

_ILLEGAL_RE = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f]")


def _decode_bin(val):
    if val is None:
        return None
    if isinstance(val, (bytes, bytearray)):
        try:
            return val.decode("euc-kr")
        except Exception:
            return val.decode("utf-8", errors="replace")
    return str(val)


def _clean(s):
    if not s:
        return s
    return _ILLEGAL_RE.sub("", s)


@bp.route("/api/affiliate/ranking")
def affiliate_ranking():
    start = request.args.get("start", "2025-01-01")
    end = request.args.get("end", "2026-03-01")
    rows = query("bar_shop1", AFFILIATE_RANKING, (start, end))
    result = []
    for r in rows:
        name = _clean(_decode_bin(r.get("company_name_bin"))) or ""
        result.append({
            "company_seq": r["company_seq"],
            "company_name": name,
            "jaehu_kind": r["JAEHU_KIND"] or "",
            "orders": r["orders"],
            "revenue": r["revenue"] or 0,
            "total_qty": r["total_qty"] or 0,
            "avg_price": r["avg_price"] or 0,
            "avg_qty": float(r["avg_qty"]) if r["avg_qty"] else 0,
        })
    return jsonify(result)


@bp.route("/api/affiliate/monthly")
def affiliate_monthly():
    start = request.args.get("start", "2025-01-01")
    end = request.args.get("end", "2026-03-01")
    rows = query("bar_shop1", AFFILIATE_MONTHLY, (start, end))
    # Build: { company_seq: { name, months: { "2025-01": {orders, revenue} } } }
    companies = {}
    for r in rows:
        seq = r["company_seq"]
        if seq not in companies:
            name = _clean(_decode_bin(r.get("company_name_bin"))) or ""
            companies[seq] = {"company_name": name, "jaehu_kind": r["JAEHU_KIND"] or "", "months": {}, "total_revenue": 0}
        m = r["month"]
        companies[seq]["months"][m] = {
            "orders": r["orders"],
            "revenue": r["revenue"] or 0,
            "total_qty": r["total_qty"] or 0,
        }
        companies[seq]["total_revenue"] += (r["revenue"] or 0)
    # Sort by total revenue desc
    sorted_list = sorted(companies.values(), key=lambda x: x["total_revenue"], reverse=True)
    return jsonify(sorted_list)


@bp.route("/api/affiliate/export")
def affiliate_export():
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    start = request.args.get("start", "2025-01-01")
    end = request.args.get("end", "2026-03-01")

    rows = query("bar_shop1", AFFILIATE_MONTHLY, (start, end))

    # Build company data
    companies = {}
    for r in rows:
        seq = r["company_seq"]
        if seq not in companies:
            name = _clean(_decode_bin(r.get("company_name_bin"))) or ""
            companies[seq] = {"name": name, "kind": r["JAEHU_KIND"] or "", "months": {},
                              "total_rev": 0, "total_ord": 0, "total_qty": 0}
        m = r["month"]
        rev = r["revenue"] or 0
        ord_ = r["orders"] or 0
        qty = r["total_qty"] or 0
        companies[seq]["months"][m] = {"orders": ord_, "revenue": rev, "qty": qty}
        companies[seq]["total_rev"] += rev
        companies[seq]["total_ord"] += ord_
        companies[seq]["total_qty"] += qty

    all_companies = list(companies.values())

    # Collect all months
    all_months = set()
    for comp in all_companies:
        all_months.update(comp["months"].keys())
    months_sorted = sorted(all_months)

    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    hdr_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    month_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    hdr_font = Font(bold=True, size=10, color="FFFFFF")
    bold_font = Font(bold=True, size=10)
    center = Alignment(horizontal="center", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")

    wb = Workbook()
    ws = wb.active
    ws.title = "제휴업체별 월별 매출"

    headers = ["순위", "업체명", "제휴종류", "주문수", "매출액", "비중", "객단가", "객수량"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = center
        cell.border = thin

    row_idx = 2

    for month in months_sorted:
        # Build rows for this month (all companies, 0 if missing)
        month_rows = []
        for comp in all_companies:
            d = comp["months"].get(month, {"orders": 0, "revenue": 0, "qty": 0})
            month_rows.append({
                "name": comp["name"], "kind": comp["kind"],
                "orders": d.get("orders", 0) or 0,
                "revenue": d.get("revenue", 0) or 0,
                "qty": d.get("qty", 0) or 0,
            })
        month_rows.sort(key=lambda x: x["revenue"], reverse=True)
        month_total_rev = sum(r["revenue"] for r in month_rows)
        month_total_ord = sum(r["orders"] for r in month_rows)

        # Month header row
        cell = ws.cell(row=row_idx, column=1, value=month)
        cell.font = bold_font
        cell.fill = month_fill
        cell.alignment = center
        for c in range(1, 9):
            ws.cell(row=row_idx, column=c).fill = month_fill
            ws.cell(row=row_idx, column=c).border = thin
        summary = f"주문 {month_total_ord:,}건 | 매출 {month_total_rev:,}원"
        cell2 = ws.cell(row=row_idx, column=2, value=summary)
        cell2.font = bold_font
        cell2.fill = month_fill
        row_idx += 1

        # Company rows
        for rank, r in enumerate(month_rows, 1):
            share = round(r["revenue"] / month_total_rev * 100, 1) if month_total_rev else 0
            avg_p = round(r["revenue"] / r["orders"]) if r["orders"] else 0
            avg_q = round(r["qty"] / r["orders"], 1) if r["orders"] else 0

            vals = [rank, r["name"], r["kind"],
                    r["orders"], r["revenue"], f"{share}%", avg_p, avg_q]
            for c, v in enumerate(vals, 1):
                cell = ws.cell(row=row_idx, column=c, value=v)
                cell.border = thin
                if c in (4, 5, 7):
                    cell.number_format = "#,##0"
                    cell.alignment = right_align
                elif c in (1, 3, 6, 8):
                    cell.alignment = center
            row_idx += 1

    # Column widths
    col_widths = {"A": 8, "B": 24, "C": 10, "D": 12, "E": 16, "F": 8, "G": 14, "H": 10}
    for letter, w in col_widths.items():
        ws.column_dimensions[letter].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:H{row_idx - 1}"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name="affiliate_monthly.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
