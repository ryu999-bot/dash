"""SKU (product catalog) list endpoint."""
from flask import Blueprint, request, jsonify, send_file
from db.connection import query
from db.queries_sku import SKU_LIST, SKU_KINDS
import io
import re

bp = Blueprint("sku", __name__)

# Regex to strip illegal XML characters for openpyxl
_ILLEGAL_RE = re.compile(
    r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f]"
)

# Card_Shape code -> label
_SHAPE_LABEL = {"1": "엽서형", "2": "접지형", "3": "폴더형", "4": "특수형"}

# CardBrand code -> label
_BRAND_LABEL = {
    "B": "바른손", "C": "더카드", "S": "비핸즈", "X": "디얼디어",
    "W": "W카드", "N": "네이처", "I": "이니스", "H": "비핸즈P",
    "F": "플라워", "D": "디자인", "P": "프리미어", "M": "모바일",
    "G": "글로벌", "U": "유니세프", "Y": "유니크", "K": "비케이",
    "T": "프리미어더카드", "A": "기타",
}

# Font type code -> label
_FONT_LABEL = {
    "a": "명조", "b": "고딕", "c": "캘리", "d": "손글씨",
    "e": "필기", "f": "둥근", "g": "장식",
}

# PrintMethod first char -> label
_PRINT_METHOD_LABEL = {
    "0": "일반", "G": "금박", "S": "은박", "I": "이리데센트",
    "Y": "유광", "N": "무광", "B": "브론즈", "F": "포일",
    "C": "컬러박", "T": "투명박", "R": "로즈골드", "D": "더블박",
    "K": "블랙박", "Z": "지문양각", "A": "아크릴", "W": "화이트박",
    "E": "엠보박", "L": "레이저박", "X": "특수박",
}


def _decode_bin(val):
    """Decode EUC-KR varbinary to string."""
    if val is None:
        return None
    if isinstance(val, (bytes, bytearray)):
        try:
            return val.decode("euc-kr")
        except Exception:
            return val.decode("utf-8", errors="replace")
    return str(val)


def _clean(s):
    """Remove illegal characters that openpyxl rejects."""
    if not s:
        return s
    return _ILLEGAL_RE.sub("", s)


# Folding code -> display label
_FOLDING_LABEL = {
    "S1": "세로1단", "S2": "세로2단", "S3": "세로3단", "S4": "세로4단",
    "G1": "가로1단", "G2": "가로2단", "G3": "가로3단",
}
# Codes where unfolded size is left blank
_FOLDING_BLANK = {"FS", "E0", "ETC", "0", "", None}


def _calc_unfolded(w, h, folding):
    """Calculate unfolded (spread) size from folding type."""
    if not w or not h or folding in _FOLDING_BLANK:
        return None, None
    code = (folding or "").strip()
    if code in ("S1", "G1"):
        return w, h
    if code == "S2":
        return w, h * 2
    if code == "S3":
        return w, h * 3
    if code == "S4":
        return w, h * 4
    if code == "G2":
        return w * 2, h
    if code == "G3":
        return w * 3, h
    return None, None


def _fetch_sku():
    # 1) Build kind map: Card_Seq -> "청첩장, 커스텀디지털카드"
    kind_rows = query("bar_shop1", SKU_KINDS)
    kind_map = {}
    for r in kind_rows:
        seq = r["Card_Seq"]
        name = _decode_bin(r.get("kind_bin")) or ""
        kind_map.setdefault(seq, []).append(name)

    # 2) Fetch products
    rows = query("bar_shop1", SKU_LIST)
    result = []
    for r in rows:
        seq = r["Card_Seq"]
        name = _decode_bin(r.get("card_name_bin"))
        kinds = ", ".join(kind_map.get(seq, []))

        w = r.get("Card_WSize") or 0
        h = r.get("Card_HSize") or 0
        folding_raw = (r.get("Card_Folding") or "").strip()
        folding_label = _FOLDING_LABEL.get(folding_raw, "")
        uw, uh = _calc_unfolded(w, h, folding_raw)

        shape_raw = (r.get("Card_Shape") or "").strip()
        shape_label = _SHAPE_LABEL.get(shape_raw, "")

        brand_raw = (r.get("CardBrand") or "").strip()
        brand_label = _BRAND_LABEL.get(brand_raw, brand_raw)

        card_material = _clean(r.get("card_material") or "") or ""
        card_group = (r.get("CARD_GROUP") or "").strip()
        group_label = "외주" if card_group == "X" else "내부" if card_group == "I" else ""

        # Post-processing (후가공)
        finishing = []
        if (r.get("IsEmbo") or "").strip() == "1":
            finishing.append("엠보")
        if (r.get("isLetterPress") or "").strip() == "1":
            finishing.append("레터프레스")
        if (r.get("isLaser") or "").strip() == "1":
            finishing.append("레이저")
        if (r.get("IsDisitalCut") or "").strip() == "1":
            finishing.append("디지털커팅")
        if (r.get("IsPunching") or "").strip() == "1":
            finishing.append("타공")
        if (r.get("isEnvSpecial") or "").strip() == "1":
            finishing.append("봉투특수")
        finishing_str = ", ".join(finishing) if finishing else ""

        # Print method (박 가공)
        pm_raw = (r.get("print_method") or "").strip()
        pm_label = ""
        if pm_raw and pm_raw != "000":
            first = pm_raw[0] if pm_raw else ""
            second = pm_raw[1] if len(pm_raw) > 1 else "0"
            pm_label = _PRINT_METHOD_LABEL.get(first, first)
            if second == "1":
                pm_label += "+양면"

        result.append({
            "card_code": r["Card_Code"],
            "card_name": _clean(name) or "",
            "card_price": r["Card_Price"] or 0,
            "product_type": r["product_type"],
            "card_kinds": _clean(kinds),
            "display": r["DISPLAY_YORN"] or "N",
            "brand": brand_label,
            "card_shape": shape_label,
            "card_material": card_material,
            "card_group": group_label,
            "finishing": finishing_str,
            "print_method": pm_label,
            "greeting_font": _clean(r.get("greeting_font") or "") or "",
            "card_wsize": w,
            "card_hsize": h,
            "card_size": f"{w} x {h}" if w and h else "",
            "folding": folding_label,
            "unfolded_wsize": uw,
            "unfolded_hsize": uh,
            "unfolded_size": f"{uw} x {uh}" if uw and uh else "",
        })
    return result


@bp.route("/api/sku/count")
def sku_count():
    rows = query("bar_shop1", "SELECT COUNT(*) AS cnt FROM S2_Card WITH (NOLOCK)")
    return jsonify({"total": rows[0]["cnt"] if rows else 0})


@bp.route("/api/sku/export")
def sku_export():
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    hdr_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    hdr_font = Font(bold=True, size=11, color="FFFFFF")
    center = Alignment(horizontal="center", vertical="center")

    rows = _fetch_sku()

    wb = Workbook()
    ws = wb.active
    ws.title = "상품목록"

    headers = ["상품코드", "SKU명", "판매가격", "브랜드", "디지털/마스터", "상품종류", "판매노출",
               "카드형태", "용지", "생산구분",
               "후가공", "박가공", "인사말폰트",
               "제품사이즈(W)", "제품사이즈(H)", "제품사이즈", "접기방식", "펼침사이즈(W)", "펼침사이즈(H)", "펼침사이즈"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = center
        cell.border = thin

    right_cols = {3, 14, 15, 18, 19}  # price, sizes
    for r_idx, row in enumerate(rows, 2):
        values = [
            row["card_code"],
            row["card_name"],
            row["card_price"],
            row["brand"],
            row["product_type"],
            row["card_kinds"],
            "Y" if row["display"] == "Y" else "N",
            row["card_shape"],
            row["card_material"],
            row["card_group"],
            row["finishing"],
            row["print_method"],
            row["greeting_font"],
            row["card_wsize"] or "",
            row["card_hsize"] or "",
            row["card_size"],
            row["folding"],
            row["unfolded_wsize"] or "",
            row["unfolded_hsize"] or "",
            row["unfolded_size"],
        ]
        for c, val in enumerate(values, 1):
            cell = ws.cell(row=r_idx, column=c, value=val)
            cell.border = thin
            if c in right_cols and isinstance(val, (int, float)):
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.alignment = center

    widths = [14, 30, 12, 10, 14, 20, 10, 10, 22, 8, 18, 12, 35, 12, 12, 14, 10, 12, 12, 14]
    for i, w in enumerate(widths, 1):
        if i <= 26:
            col_letter = chr(64 + i)
        else:
            col_letter = "A" + chr(64 + i - 26)
        ws.column_dimensions[col_letter].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:T{len(rows) + 1}"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return send_file(buf, as_attachment=True, download_name="sku_list.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
