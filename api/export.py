"""Excel export endpoint."""
import io
from flask import Blueprint, request, send_file
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from db.connection import query
from db import queries_bar_shop1 as q1
from api.kpi import _parse_dates, _parse_basis

bp = Blueprint("export", __name__)

THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
CENTER = Alignment(horizontal="center", vertical="center")
NUM_FMT = "#,##0"


@bp.route("/api/export")
def export():
    start, end = _parse_dates(request.args)
    basis = _parse_basis(request.args)
    rows = query("bar_shop1", q1.detail_daily(basis), (start, end))

    wb = Workbook()
    ws = wb.active
    ws.title = "상세데이터"

    headers = ["날짜", "사이트", "상품구분", "주문수", "매출액", "객단가"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    for r_idx, row in enumerate(rows, 2):
        aov = round(row["revenue"] / row["order_count"]) if row["order_count"] else 0
        values = [
            row["order_day"],
            row["site_name"],
            row["product_type"],
            row["order_count"],
            row["revenue"] or 0,
            aov,
        ]
        for c, val in enumerate(values, 1):
            cell = ws.cell(row=r_idx, column=c, value=val)
            cell.border = THIN_BORDER
            if isinstance(val, (int, float)) and c >= 4:
                cell.number_format = NUM_FMT
                cell.alignment = Alignment(horizontal="right")
            elif c <= 3:
                cell.alignment = CENTER

    # Auto-width
    widths = [12, 14, 12, 10, 14, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"sales_detail_{start}_{end}.xlsx"
    return send_file(buf, as_attachment=True, download_name=filename,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
