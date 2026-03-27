"""Email sending API - upload Excel, compose, attach files, send via Gmail SMTP."""
import os
import smtplib
import tempfile
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

from flask import Blueprint, request, jsonify
from openpyxl import load_workbook
from werkzeug.utils import secure_filename

bp = Blueprint("mail", __name__)

# Temp storage for uploaded data
_recipients = []  # [{company, email}, ...]
_attachments_dir = tempfile.mkdtemp(prefix="mail_attach_")


def _get_gmail_config():
    """Read Gmail credentials from env."""
    user = os.getenv("GMAIL_USER", "")
    password = os.getenv("GMAIL_APP_PASSWORD", "")
    return user, password


@bp.route("/api/mail/upload", methods=["POST"])
def upload_excel():
    """Parse uploaded Excel and return recipient list."""
    global _recipients

    f = request.files.get("file")
    if not f:
        return jsonify({"error": "파일이 없습니다."}), 400

    fname = secure_filename(f.filename or "upload.xlsx")
    if not fname.endswith((".xlsx", ".xls")):
        return jsonify({"error": "엑셀 파일(.xlsx)만 지원합니다."}), 400

    # Save temp and parse
    tmp_path = os.path.join(tempfile.gettempdir(), fname)
    f.save(tmp_path)

    try:
        wb = load_workbook(tmp_path, read_only=True)
        ws = wb.active
        recipients = []

        for row in ws.iter_rows(min_row=1, values_only=True):
            if not row or len(row) < 2:
                continue
            company = str(row[0] or "").strip()
            email = str(row[1] or "").strip()
            if not company or not email or "@" not in email:
                continue
            recipients.append({"company": company, "email": email})
        wb.close()
    except Exception as e:
        return jsonify({"error": f"엑셀 파싱 실패: {str(e)}"}), 400
    finally:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)

    _recipients = recipients
    return jsonify({"count": len(recipients), "recipients": recipients})


@bp.route("/api/mail/attach", methods=["POST"])
def upload_attachment():
    """Upload attachment file."""
    f = request.files.get("file")
    if not f:
        return jsonify({"error": "파일이 없습니다."}), 400

    fname = secure_filename(f.filename or "attachment")
    save_path = os.path.join(_attachments_dir, fname)
    f.save(save_path)

    return jsonify({"filename": fname, "path": save_path})


@bp.route("/api/mail/attach/clear", methods=["POST"])
def clear_attachments():
    """Clear all uploaded attachments."""
    for fn in os.listdir(_attachments_dir):
        fp = os.path.join(_attachments_dir, fn)
        if os.path.isfile(fp):
            os.remove(fp)
    return jsonify({"ok": True})


@bp.route("/api/mail/send", methods=["POST"])
def send_mail():
    """Send emails to all recipients."""
    data = request.json or {}
    subject = data.get("subject", "").strip()
    body = data.get("body", "").strip()
    selected = data.get("recipients", [])  # list of email strings, or empty=all

    if not subject:
        return jsonify({"error": "제목을 입력하세요."}), 400
    if not body:
        return jsonify({"error": "본문을 입력하세요."}), 400

    gmail_user, gmail_pass = _get_gmail_config()
    if not gmail_user or not gmail_pass:
        return jsonify({"error": "Gmail 설정이 없습니다. .env에 GMAIL_USER, GMAIL_APP_PASSWORD를 추가하세요."}), 400

    # Determine recipients
    if selected:
        targets = [r for r in _recipients if r["email"] in selected]
    else:
        targets = _recipients

    if not targets:
        return jsonify({"error": "수신자가 없습니다. 엑셀을 먼저 업로드하세요."}), 400

    # Collect attachment files
    attach_files = []
    for fn in os.listdir(_attachments_dir):
        fp = os.path.join(_attachments_dir, fn)
        if os.path.isfile(fp):
            attach_files.append((fn, fp))

    # Send
    results = {"success": 0, "fail": 0, "errors": []}
    try:
        server = smtplib.SMTP("smtp.gmail.com", 587)
        server.starttls()
        server.login(gmail_user, gmail_pass)

        for r in targets:
            try:
                msg = MIMEMultipart()
                msg["From"] = gmail_user
                msg["To"] = r["email"]
                # Personalize subject/body with company name
                personal_subject = subject.replace("{업체명}", r["company"])
                personal_body = body.replace("{업체명}", r["company"])
                msg["Subject"] = personal_subject
                msg.attach(MIMEText(personal_body, "plain", "utf-8"))

                # Attach files
                for fname, fpath in attach_files:
                    with open(fpath, "rb") as af:
                        part = MIMEBase("application", "octet-stream")
                        part.set_payload(af.read())
                    encoders.encode_base64(part)
                    part.add_header("Content-Disposition", f"attachment; filename=\"{fname}\"")
                    msg.attach(part)

                server.send_message(msg)
                results["success"] += 1
            except Exception as e:
                results["fail"] += 1
                results["errors"].append({"email": r["email"], "error": str(e)})

        server.quit()
    except Exception as e:
        return jsonify({"error": f"SMTP 연결 실패: {str(e)}"}), 500

    # Cleanup attachments after send
    for fn in os.listdir(_attachments_dir):
        fp = os.path.join(_attachments_dir, fn)
        if os.path.isfile(fp):
            os.remove(fp)

    return jsonify(results)
